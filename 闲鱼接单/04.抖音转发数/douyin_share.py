# from DrissionPage import SessionPage
import sys
import random
from os import path
import pyperclip
import openpyxl
import tkinter as tk
import tkinter.messagebox
from time import sleep
from tkinter import ttk
from DrissionPage import ChromiumPage
from DrissionPage import ChromiumPage, ChromiumOptions
from DrissionPage.common import By

dict_data = {}
# co = ChromiumOptions().headless()
page = ChromiumPage()

work_book = openpyxl.load_workbook(r"D:\数据统计话题&视频.xlsx")
# 获取xlsx文件的所有工作表
sheets = work_book.sheetnames

num = 0
all_num = 0

# 选择第一个工作表
for i in range(len(sheets)):
  url_data = []
  sheet = work_book[sheets[i]]
  # 选择要读取的列
  column_index = 3
  # 获取url数据
  for row in sheet.rows:
    url_data.append(row[column_index-1].value)
  url_data.remove('视频链接')

  # 得到并写入
  for j in range(len(url_data)):
    if url_data[j] == None and url_data[j+1] == None:
      break
    try:
      # not_exist = ''
      page.get(url_data[j].strip())
      page.wait(3)
      # if not_exist == '你要观看的视频不存在':
      #   print(f'换代理')
      #   page.wait(60)
      try:
        # S1XIwUxW
        big_div = page.ele('.dkgrBha5')
        # big_div_tow = big_div.child(2)
        lens = len(big_div.children)
        share = big_div.child(6)
        share1 = share.child(1)
        share2 = share1.child(2)
        # share = page.ele('x:/html/body/div[3]/div[1]/div[4]/div[2]/div/div/div[1]/div[2]/div/div[2]/div[2]/div[1]/div[1]/div[2]')
        share = page.ele('.JSdgvKUi PkJ9Apgu vV5XGmMB').text
        if share == '分享':
          share = 0
      except:
        try:
          share_index = page.eles('.G0CbEcWs')
          share = share_index[3].text
        except Exception as e:
            share = '错误！'


      # 将share添加进xlsx文件D列
      sheet.cell(row=j+2, column=4).value = share
      # 进度
      num += 1
    except Exception as e:
      print(f'第{j+1}个url出错！')
      print(f'错误信息：{e}')

    print(f'当前进度：{j+1}/{len(url_data)}  当前的share数为{share}')
    # 完成一个表 写入一个表
  work_book.save(r"D:\数据统计话题&视频.xlsx")
    # 进度
  all_num += len(url_data)
  print(f'第{i+1}个表写入完成！')
  print(f'当前共计{all_num}个视频的share数')
page.quit()
 
# 关闭工作簿
work_book.close()

print('数据统计完成！')
